import time,os,datetime

import openpyxl
from openpyxl import Workbook, load_workbook

from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from magicdata_setup.config import config
from magicdata_setup import randomeString
from log_config import setup_logger
from constant import validation_assert,creds,error,input_field
from magicdata_setup.randomeString import generate_random_emoji_string

logger = setup_logger()
chrome_options = webdriver.ChromeOptions()
driver = webdriver.Chrome(options=chrome_options)
chrome_options.add_argument('--headless')
driver.maximize_window()
logger.info("Opening signup page")
driver.get(config.WEB_URL)
email = config.TEAM_EMAIL
password = config.CORRECT_PASSWORD
confirm_password = config.CONFIRM_PASSWORD
wait = WebDriverWait(driver, 60)
clear_input = Keys.SPACE+ Keys.CONTROL + 'a' + Keys.BACKSPACE
signup_full_name = creds.VALID_FULL_NAME


def create_an_account_btn():
    return wait.until(EC.element_to_be_clickable((By.XPATH, "//button[normalize-space()='Create Account']")))

def full_name_input_field():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder='First Name']")))

def full_name_length_validation():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Full name is too long')]")))

def email_length_validation():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Email is too long')]")))

def full_name_special_validation():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Full name should contain only letters and numbers')]")))

def profile_email_input_field():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Email']")))

def login_email_input_field():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Enter your email address']")))

def password_input_field():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Enter your password']")))

def confirm_password_input_field():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Confirm your password']")))

def signup_btn():

    return wait.until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Sign Up']")))

def save_btn():

    return wait.until(EC.element_to_be_clickable((By.XPATH, "//button[normalize-space()='Save']")))

def check_blank_fullname():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Full name is required')]")))

def check_blank_email():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Email is required')]")))

def check_blank_password():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Password is required')]")))

def username_exists_validation():
    return wait.until(EC.visibility_of_element_located((
        By.XPATH, "//span[contains(text(),'Username taken. Please choose another')]"
    )))


def exist_email_validation():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'The email has already been taken.')]")))

def email_invalid_validation():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'Email address is not valid.')]")))

def check_password_length_validation():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'The Password field must be at least 8 characters')]")))

def check_strong_password_validation():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'The Password Must include uppercase, lowercase, number, and special character')]")))

def success_signup_message():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//p[contains(text(),'Congratulations! Your new account has been successfully created!')]")))

def refresh_page():
    logger.info("Refreshing page")
    time.sleep(1)
    return driver.refresh()

def quit_browser():
    logger.info("Quitting browser")
    return driver.quit()

def overlay_spinner():
    return WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, "overlay-spinner")))

def save_screenshot(filename, use_timestamp=True, folder="screenshorts"):
    os.makedirs(folder, exist_ok=True)

    if use_timestamp:
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = "{}_{}.png".format(filename, timestamp)
    else:
        filename = "{}.png".format(filename)

    full_path = folder, filename
    driver.save_screenshot(f"{folder}/{filename}")
    return full_path


def valid_email_validation():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//p[contains(text(),'Invalid email(s):')]")))




def password_mismatch_validation():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),\"Confirm Password doesn't match with Password.\")]")))

def password_pattern_validation():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Password must contain at least: 1 uppercase letter,1 lowercase letter, 1 number and 1 special character.')]")))

def signup_success_message():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Signup successful')]")))

def existing_user_error_message():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'Email already registered')]")))

def toggle_visibility_icon():
    return wait.until(EC.element_to_be_clickable((By.XPATH, "//span[@role='button']")))

def signin_btn():
    return wait.until(EC.element_to_be_clickable((By.XPATH, "//button[normalize-space()='Sign In']")))

def check_profile():
    return wait.until(EC.element_to_be_clickable((By.XPATH, "//span[contains(text(),'Profile')]")))

def check_team():
    return wait.until(EC.element_to_be_clickable((By.XPATH, "//span[contains(text(),'Team')]")))

def check_security():
    return wait.until(EC.element_to_be_clickable((By.XPATH, "//span[normalize-space()='Security']")))

def check_success_profile_update():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'User updated successfully')]")))

def check_invite_success_msg():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'Invitation sent successfully')]")))

def get_updated_full_name():
    return wait.until(EC.presence_of_element_located((
        By.NAME, "fullName"
    )))


def avtar_icon():
    return wait.until(EC.element_to_be_clickable((By.XPATH, "//span[contains(@class,'avatar-icon')]//*[name()='svg']")))

def sign_out():
    return wait.until(EC.element_to_be_clickable((By.XPATH, "//span[normalize-space()='Sign Out']")))

def sign_out_btn():
    return wait.until(EC.element_to_be_clickable((By.XPATH, "//button[normalize-space()='Sign Out']")))

def verify_avtar_email():
    return wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='text-xs']")))

def verify_avtar_full_name():
    return wait.until(EC.element_to_be_clickable((By.XPATH, "//div[contains(@class, 'font-bold') and contains(@class, 'text-gray-900')]")))

def invite_btn():
    time.sleep(1)

    return wait.until(EC.element_to_be_clickable((By.XPATH, "//button[normalize-space()='Invite Teammates']")))

def invite_email_input_field():
    return wait.until(
        EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Add emails separated by commas (e.g., user1@example.com, user2@example.com)']")))

def send_invite_btn():
    return wait.until(EC.element_to_be_clickable((By.XPATH, "//button[normalize-space()='Send Invite']")))

def close_alert():
    close_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@type='button']")))
    action = ActionChains(driver)
    action.double_click(close_button).perform()

def get_multiple_invite_emails(filepath="signup_credentials.xlsx", sheetname="Users"):
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheetname]

    emails = []
    for row in ws.iter_rows(min_row=2, max_row=9, min_col=2, max_col=2):
        for cell in row:
            if cell.value:
                emails.append(cell.value.strip())

    return ",".join(emails)


# ============================== TEST CASES ==============================

class TestAddTeam:

    def test_valid_login_flow(self):
        logger.info("Running test: Valid login flow")
        refresh_page()
        login_email_input_field().send_keys(config.TEAM_EMAIL)
        password_input_field().send_keys(config.CORRECT_PASSWORD)
        toggle_visibility_icon().click()
        logger.debug("Clicked password visibility toggle")
        signin_btn().click()
        time.sleep(5)
        assert validation_assert.DASHBOARD in driver.current_url
        logger.info("Signup successful and redirected to dashboard/home")

    def test_my_teams(self):
        logger.info("Navigating to Teams")
        check_team().click()

    def test_invite_single_user(self):
        logger.info("Inviting single team member")
        invite_btn().click()
        invite_email_input_field().send_keys(input_field.VALID_SINGLE_EMAIL)
        send_invite_btn().click()
        assert check_invite_success_msg().text== validation_assert.INVITE_SUCCESS_MSG

    def test_invite_multiple_user(self):
        logger.info("Inviting multiple team members")
        invite_btn().click()
        invite_email_input_field().send_keys(input_field.VALID_MULTIPLE_EMAIL)
        send_invite_btn().click()
        assert check_invite_success_msg().text== validation_assert.INVITE_SUCCESS_MSG


    def test_invite_trimmed_email(self):
        logger.info("Testing invite with trimmed email")
        WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.CLASS_NAME, "dialog-overlay")))
        invite_btn().click()
        invite_email_input_field().send_keys("  "+ input_field.VALID_SINGLE_EMAIL +"  ")  # with leading/trailing space
        send_invite_btn().click()
        assert check_invite_success_msg().text == validation_assert.INVITE_SUCCESS_MSG

    def test_invite_mixed_case_email(self):
        logger.info("Testing invite with mixed case email")

        invite_btn().click()
        invite_email_input_field().send_keys(input_field.MIXED_EMAIL_CASE)
        send_invite_btn().click()
        assert check_invite_success_msg().text == validation_assert.INVITE_SUCCESS_MSG


    def test_invalid_email_format(self):
        logger.info("Testing invalid email format")
        invite_btn().click()
        invite_email_input_field().send_keys(input_field.INVALID_EMAIL)
        send_invite_btn().click()
        assert error.ENTER_VALID_TEAM_EMAIL in valid_email_validation().text
        close_alert()



    def test_multiple_invalid_emails(self):
        logger.info("Testing multiple invalid emails")
        invite_btn().click()
        invite_email_input_field().send_keys(input_field.MULTIPLE_INVALID_EMAIL)
        send_invite_btn().click()
        assert error.ENTER_VALID_TEAM_EMAIL in valid_email_validation().text
        close_alert()




    def test_already_invited_email(self):
        logger.info("Testing already invited email scenario")
        invite_btn().click()
        invite_email_input_field().send_keys(input_field.ALREADY_INVITED_EMAIL)
        send_invite_btn().click()
        assert check_invite_success_msg().text == validation_assert.INVITE_SUCCESS_MSG

    def test_invite_multiple_user_through_excel(self):
        logger.info("Inviting multiple team members from Excel sheet")
        emails = get_multiple_invite_emails()  # Read from Excel
        assert emails, "No emails fetched from Excel"
        invite_btn().click()
        invite_email_input_field().send_keys(emails)
        send_invite_btn().click()
        assert check_invite_success_msg().text == validation_assert.INVITE_SUCCESS_MSG
        logger.info("Multiple users invited successfully")

