import time,os,datetime
from openpyxl import Workbook, load_workbook

from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from magicdata_setup.config import config
from magicdata_setup import randomeString
from log_config import setup_logger
from constant import validation_assert,creds,error
logger = setup_logger()
chrome_options = webdriver.ChromeOptions()
driver = webdriver.Chrome(options=chrome_options)
chrome_options.add_argument('--headless')
driver.maximize_window()
logger.info("Opening signup page")
driver.get(config.WEB_URL)
email = config.CORRECT_EMAIL
password = config.CORRECT_PASSWORD
confirm_password = config.CORRECT_PASSWORD
email_signup = creds.VALID_UNIQUE_EMAIL
full_name = creds.VALID_FULL_NAME

wait = WebDriverWait(driver, 60)
clear_input = Keys.SPACE+ Keys.CONTROL + 'a' + Keys.BACKSPACE



def create_an_account_btn():
    return wait.until(EC.element_to_be_clickable((By.XPATH, "//button[normalize-space()='Create Account']")))

def full_name_input_field():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Enter your full name']")))

def full_name_length_validation():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Full Name is too long')]")))

def email_length_validation():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Email is too long')]")))

def full_name_special_validation():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Full Name should contain only letters and numbers')]")))

def email_input_field():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Enter your email address']")))

def password_input_field():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Enter your password']")))

def confirm_password_input_field():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Confirm your password']")))

def signup_btn():

    return wait.until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Sign Up']")))

def check_blank_fullname():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Please enter your full name')]")))

def check_blank_email():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Email is required')]")))

def check_blank_password():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Password is required')]")))

def username_exists_validation():
    return wait.until(EC.visibility_of_element_located((
        By.XPATH, "//span[contains(text(),'Username taken. Please choose another')]"
    )))


def exist_email_validation():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'The email has already been taken.')]")))

def email_invalid_validation():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'Email address is not valid.')]")))

def check_password_length_validation():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'The Password field must be at least 8 characters')]")))

def check_strong_password_validation():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'The Password Must include uppercase, lowercase, number, and special character')]")))

def success_signup_message():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//p[contains(text(),'Congratulations! Your new account has been successfully created!')]")))

def refresh_page():
    logger.info("Refreshing page")
    time.sleep(1)
    return driver.refresh()

def quit_browser():
    logger.info("Quitting browser")
    return driver.quit()

def overlay_spinner():
    return WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, "overlay-spinner")))

def save_screenshot(filename, use_timestamp=True, folder="screenshorts"):
    os.makedirs(folder, exist_ok=True)

    if use_timestamp:
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = "{}_{}.png".format(filename, timestamp)
    else:
        filename = "{}.png".format(filename)

    full_path = folder, filename
    driver.save_screenshot(f"{folder}/{filename}")
    return full_path

def valid_email_validation():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Please enter a valid email address')]")))  # Adjust text if needed

def password_mismatch_validation():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),\"Confirm Password doesn't match with Password.\")]")))

def password_pattern_validation():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Password must contain at least: 1 uppercase letter,1 lowercase letter, 1 number and 1 special character.')]")))

def signup_success_message():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Signup successful')]")))

def existing_user_error_message():
    return wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'Email already registered')]")))

def toggle_visibility_icon():
    return wait.until(EC.element_to_be_clickable((By.XPATH, "//span[@role='button']")))

def signin_btn():
    return wait.until(EC.element_to_be_clickable((By.XPATH, "//button[normalize-space()='Sign In']")))

def avtar_icon():
    return wait.until(EC.element_to_be_clickable((By.XPATH, "//span[contains(@class,'avatar-icon')]//*[name()='svg']")))
def verify_avtar_email():
    return wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='text-xs']")))




def write_credentials_to_excel(fullname,email, password,confirm_password, filename="signup_credentials.xlsx", sheet_name="Users"):
    try:
        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb[sheet_name]
        else:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(["Full Name","Email", "Password","Confirm password"])

        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=1).value = fullname
        ws.cell(row=next_row, column=2).value = email
        ws.cell(row=next_row, column=3).value = password
        ws.cell(row=next_row, column=4).value = confirm_password

        wb.save(filename)
        print(f"Saved in: {filename}, Row: {next_row}")
    except Exception as e:
        print(f"Failed to write to Excel: {e}")

# ============================== TEST CASES ==============================

class TestSignup:

    def test_blank_field_validation(self):
        logger.info("Running test: Blank field validation")
        signup_btn().click()
        full_name_input_field().send_keys(clear_input)
        assert check_blank_fullname().text == validation_assert.ENTER_FULL_NAME
        logger.info("Blank full name validation passed")
        email_input_field().send_keys('/'+clear_input)
        assert check_blank_email().text == validation_assert.ENTER_SIGNUP_EMAIL
        logger.info("Blank email validation passed")
        password_input_field().send_keys(clear_input)
        assert check_blank_password().text == validation_assert.ENTER_SIGNUP_PASSWORD
        logger.info("Blank password validation passed")

    def test_full_name_length_validation(self):
        logger.info("Running test: Full name max length validation")
        refresh_page()
        full_name_input_field().send_keys(creds.LONG_FULL_NAME)
        email_input_field().send_keys(email)
        password_input_field().send_keys(password)
        confirm_password_input_field().send_keys(confirm_password)
        create_an_account_btn().click()
        assert full_name_length_validation().text == error.LENGTH_FULL_NAME_VALIDATION
        logger.info("Full name length validation passed")


    def test_full_name_special_validation(self):
        logger.info("Running test: Full name special character validation")
        refresh_page()
        full_name_input_field().send_keys(creds.INVALID_FULL_NAME)
        email_input_field().send_keys(email)
        password_input_field().send_keys(password)
        confirm_password_input_field().send_keys(confirm_password)
        create_an_account_btn().click()
        assert full_name_special_validation().text == error.ENTER_VALID_FULLNAME
        logger.info("Full name special character validation passed")


    def test_email_case_insensitivity(self):
        logger.info("Running test: Email case insensitivity validation")
        refresh_page()
        full_name_input_field().send_keys(creds.VALID_FULL_NAME)
        email_input_field().send_keys(email.upper())
        password_input_field().send_keys(password)
        confirm_password_input_field().send_keys(password)
        create_an_account_btn().click()
        assert existing_user_error_message().text == error.EMAIL_ALREADY_EXISTS
        logger.info("Email case-insensitivity handled correctly")
    def test_invalid_email_format(self):
        logger.info("Running test: Invalid email format validation")
        refresh_page()
        full_name_input_field().send_keys(creds.VALID_FULL_NAME)
        email_input_field().send_keys(creds.INVALID_EMAIL_FORMAT)
        password_input_field().send_keys(password)
        confirm_password_input_field().send_keys(confirm_password)
        create_an_account_btn().click()
        assert valid_email_validation().text == error.ENTER_VALID_EMAIL
        logger.info("Invalid email format validation passed")


    def test_very_long_email(self):
        logger.info("Running test: Very long email address")
        refresh_page()
        full_name_input_field().send_keys(creds.VALID_FULL_NAME)
        email_input_field().send_keys(creds.LONG_EMAIL)
        password_input_field().send_keys(password)
        confirm_password_input_field().send_keys(password)
        create_an_account_btn().click()
        assert email_length_validation().text == error.LENGTH_EMAIL_VALIDATION
        logger.info("Long email validation passed")
    def test_password_pattern_validation(self):
        logger.info("Running test: Password pattern validation")
        refresh_page()
        full_name_input_field().send_keys(creds.VALID_FULL_NAME)
        email_input_field().send_keys(email)
        password_input_field().send_keys(creds.WEAK_PASSWORD)
        confirm_password_input_field().send_keys(creds.WEAK_PASSWORD)
        create_an_account_btn().click()
        assert password_pattern_validation().text == error.PASSWORD_PATTERN_VALIDATION
        logger.info("Password pattern validation passed")

    def test_password_case_sensitivity(self):
        logger.info("Running test: Password case sensitivity")
        refresh_page()
        full_name_input_field().send_keys(creds.VALID_FULL_NAME)
        email_input_field().send_keys(email_signup)
        password_input_field().send_keys("Test@123")
        confirm_password_input_field().send_keys("test@123")
        create_an_account_btn().click()
        assert password_mismatch_validation().text == error.PASSWORD_NOT_MATCHING
        logger.info("Password case mismatch correctly validated")

    def test_password_mismatch_validation(self):
        logger.info("Running test: Password mismatch validation")
        refresh_page()
        full_name_input_field().send_keys(creds.VALID_FULL_NAME)
        email_input_field().send_keys(email)
        password_input_field().send_keys(creds.PASSWORD)
        confirm_password_input_field().send_keys(creds.MISMATCH_PASSWORD)
        create_an_account_btn().click()
        assert password_mismatch_validation().text == error.PASSWORD_NOT_MATCHING
        logger.info("Password mismatch validation passed")

    def test_existing_user_signup(self):
        logger.info("Running test: Existing user signup validation")
        refresh_page()

        full_name_input_field().send_keys(creds.VALID_FULL_NAME)
        email_input_field().send_keys(email)
        password_input_field().send_keys(password)
        confirm_password_input_field().send_keys(confirm_password)
        create_an_account_btn().click()
        assert existing_user_error_message().text == error.EMAIL_ALREADY_EXISTS
        logger.info("Signup blocked for existing user – validation passed")


    def test_successful_signup(self):
        logger.info("Running test: Successful signup flow")
        refresh_page()
        full_name_input_field().send_keys(creds.VALID_FULL_NAME)
        email_input_field().send_keys(email_signup)
        print(email_signup)
        password_input_field().send_keys(password)
        confirm_password_input_field().send_keys(confirm_password)
        create_an_account_btn().click()
        time.sleep(3)
        assert validation_assert.DASHBOARD in driver.current_url
        logger.info("Signup successful and redirected to dashboard/home")
        write_credentials_to_excel(full_name,email_signup,password,confirm_password)

    def test_valid_user_signup(self):
        logger.info("Running test: Valid login flow")
        refresh_page()
        avtar_icon().click()
        time.sleep(3)
        assert verify_avtar_email().text == email_signup
        quit_browser()



